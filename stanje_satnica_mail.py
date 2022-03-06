import openpyxl
import smtplib
from email.message import EmailMessage


EXCEL_FILE_LOCATION = None
WORKBOOK_NAME = None
WORKSHEET_NAME = None
DISPATCHER_EMAIL = None
SUBJECT = None
TEMPLATE = """
Stanje satnice zaključno sa prošlim mjesecom:

Slobodni sati: {ss},
Preostali godišnji odmor: {gg} dana,
Prošli mjesec odrađeno: {os},
Promjena u fondu sati: {razlika}
"""



def send_mail(employee, s):
    msg = EmailMessage()
    msg.set_content(TEMPLATE.format(*employee))
    msg['Subject'] = SUBJECT
    msg['From'] = DISPATCHER_EMAIL
    msg['To'] = employee['email']
    s.send_message(msg)


if __name__ == '__main__':
    s = smtplib.SMTP('localhost')
    wb = openpyxl.load_workbook(EXCEL_FILE_LOCATION + WORKBOOK_NAME)
    ws = wb.get_sheet_by_name(WORKSHEET_NAME)
    col_names = [_col[0].value for _col in ws.iter_cols]
    for row in ws.iter_rows(min_row=2):
        employee = dict(zip(col_names, row))
        send_mail(employee, s)
    s.quit()







